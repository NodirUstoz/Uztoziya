import cv2
import numpy as np
import pytesseract
from PIL import Image
import re
import json
import logging
import os
import io
from django.conf import settings
from google.cloud import vision
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .models import OCRProcessing, TestResult

logger = logging.getLogger(__name__)


class OCRService:
    """OCR xizmati"""
    
    def __init__(self):
        # Tesseract yo'lini sozlash (Windows uchun)
        if hasattr(settings, 'TESSERACT_PATH'):
            pytesseract.pytesseract.tesseract_cmd = settings.TESSERACT_PATH
        
        # Google Vision API'ni sozlash
        try:
            # API key orqali autentifikatsiya
            import os
            os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = getattr(settings, 'GOOGLE_API_KEY', '')
            self.vision_client = vision.ImageAnnotatorClient()
        except Exception as e:
            logger.warning(f"Google Vision API sozlanmadi: {e}")
            self.vision_client = None
    
    def preprocess_image(self, image_path):
        """Rasmni oldindan qayta ishlash"""
        try:
            # Rasmni o'qish
            image = cv2.imread(image_path)
            
            # Rangli rasmni kulrangga aylantirish
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Shovqinni kamaytirish
            denoised = cv2.medianBlur(gray, 3)
            
            # Kontrastni oshirish
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            enhanced = clahe.apply(denoised)
            
            # Otsu thresholding
            _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            return thresh
        except Exception as e:
            logger.error(f"Rasmni qayta ishlashda xatolik: {e}")
            return None
    
    def extract_text(self, image_path):
        """Rasmdan matnni ajratib olish"""
        try:
            # Avval Google Vision API'ni sinab ko'rish
            if self.vision_client:
                text, confidence = self.extract_text_google(image_path)
                if text and confidence > 0.7:  # Google Vision natijasi yaxshi bo'lsa
                    return text, confidence
            
            # Google Vision ishlamasa yoki natija yomon bo'lsa, Tesseract ishlatish
            return self.extract_text_tesseract(image_path)
            
        except Exception as e:
            logger.error(f"OCR qilishda xatolik: {e}")
            return None, 0.0
    
    def extract_text_google(self, image_path):
        """Google Vision API orqali OCR"""
        try:
            with io.open(image_path, 'rb') as image_file:
                content = image_file.read()
            
            image = vision.Image(content=content)
            response = self.vision_client.text_detection(image=image)
            texts = response.text_annotations
            
            if texts:
                # Birinchi text - butun rasm matni
                full_text = texts[0].description
                # Ishonch darajasini hisoblash
                confidence = 0.9  # Google Vision odatda yuqori ishonch darajasi
                return full_text.strip(), confidence
            
            return None, 0.0
            
        except Exception as e:
            logger.error(f"Google Vision OCR xatoligi: {e}")
            return None, 0.0
    
    def extract_text_tesseract(self, image_path):
        """Tesseract orqali OCR"""
        try:
            # Rasmni qayta ishlash
            processed_image = self.preprocess_image(image_path)
            
            if processed_image is None:
                return None, 0.0
            
            # OCR qilish
            custom_config = r'--oem 3 --psm 6 -l uzb+eng'
            text = pytesseract.image_to_string(processed_image, config=custom_config)
            
            # Ishonch darajasini olish
            data = pytesseract.image_to_data(processed_image, output_type=pytesseract.Output.DICT)
            confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
            avg_confidence = sum(confidences) / len(confidences) if confidences else 0
            
            return text.strip(), avg_confidence
            
        except Exception as e:
            logger.error(f"Tesseract OCR xatoligi: {e}")
            return None, 0.0
    
    def parse_test_answers(self, text):
        """Test javoblarini tahlil qilish"""
        try:
            # O'quvchi ismini topish
            student_name = self.extract_student_name(text)
            
            # Javoblarni topish
            answers = self.extract_answers(text)
            
            return {
                'student_name': student_name,
                'answers': answers
            }
        except Exception as e:
            logger.error(f"Test javoblarini tahlil qilishda xatolik: {e}")
            return None
    
    def extract_student_name(self, text):
        """O'quvchi ismini ajratib olish"""
        # Ism uchun regex patternlar
        name_patterns = [
            r'ism[:\s]*([A-Za-z\u0400-\u04FF\u0500-\u052F\u2D00-\u2D2F\u2D30-\u2D7F\s]+)',
            r'foydalanuvchi[:\s]*([A-Za-z\u0400-\u04FF\u0500-\u052F\u2D00-\u2D2F\u2D30-\u2D7F\s]+)',
            r'ismi[:\s]*([A-Za-z\u0400-\u04FF\u0500-\u052F\u2D00-\u2D2F\u2D30-\u2D7F\s]+)',
        ]
        
        for pattern in name_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        
        return "Noma'lum o'quvchi"
    
    def extract_answers(self, text):
        """Javoblarni ajratib olish"""
        answers = {}
        
        # Javob uchun regex patternlar
        answer_patterns = [
            r'(\d+)[\.\)]\s*([A-Da-d])',  # 1. A, 2) B format
            r'savol\s*(\d+)[:\s]*([A-Da-d])',  # Savol 1: A format
            r'(\d+)\s*-\s*([A-Da-d])',  # 1 - A format
        ]
        
        for pattern in answer_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for question_num, answer in matches:
                answers[int(question_num)] = answer.upper()
        
        return answers


class TestGradingService:
    """Test baholash xizmati"""
    
    def __init__(self):
        self.ocr_service = OCRService()
    
    def grade_test(self, ocr_processing, test):
        """Testni baholash"""
        try:
            # OCR natijasini olish
            if not ocr_processing.processed_text:
                return None
            
            # Javoblarni tahlil qilish
            parsed_data = self.ocr_service.parse_test_answers(ocr_processing.processed_text)
            
            if not parsed_data:
                return None
            
            student_name = parsed_data['student_name']
            student_answers = parsed_data['answers']
            
            # Test savollarini olish
            questions = test.questions.all().order_by('order')
            
            total_questions = questions.count()
            correct_answers = 0
            wrong_answers = 0
            
            # Har bir savolni tekshirish
            for question in questions:
                question_num = question.order
                student_answer = student_answers.get(question_num)
                
                if student_answer:
                    # To'g'ri javobni topish
                    correct_answer = question.answers.filter(is_correct=True).first()
                    
                    if correct_answer and student_answer == correct_answer.answer_text:
                        correct_answers += 1
                    else:
                        wrong_answers += 1
            
            # Ball va foizni hisoblash
            score = correct_answers
            percentage = (correct_answers / total_questions * 100) if total_questions > 0 else 0
            
            # Baholash
            grade = self.calculate_grade(percentage)
            
            # Natijani saqlash
            test_result = TestResult.objects.create(
                ocr_processing=ocr_processing,
                student_name=student_name,
                total_questions=total_questions,
                correct_answers=correct_answers,
                wrong_answers=wrong_answers,
                score=score,
                percentage=percentage,
                grade=grade
            )
            
            return test_result
            
        except Exception as e:
            logger.error(f"Test baholashda xatolik: {e}")
            return None
    
    def calculate_grade(self, percentage):
        """Baholash hisoblash"""
        if percentage >= 90:
            return "A'lo"
        elif percentage >= 80:
            return "Yaxshi"
        elif percentage >= 70:
            return "Qoniqarli"
        elif percentage >= 60:
            return "Qoniqarsiz"
        else:
            return "Yomon"


class ExcelExportService:
    """Excel eksport xizmati"""
    
    def __init__(self):
        self.header_font = Font(bold=True)
    
    def _auto_fit_columns(self, worksheet):
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    max_length = max(max_length, cell_length)
                except Exception:
                    continue
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    
    def export_test_results(self, test, results):
        """Test natijalarini Excel ga eksport qilish"""
        try:
            if not results:
                logger.warning("Excel eksport qilish uchun natijalar topilmadi")
                return None
            
            export_dir = os.path.join('media', 'excel_exports')
            os.makedirs(export_dir, exist_ok=True)
            timestamp = results[0].processed_at.strftime('%Y%m%d_%H%M%S')
            excel_path = os.path.join(export_dir, f"test_{test.id}_{timestamp}.xlsx")
            
            workbook = Workbook()
            results_sheet = workbook.active
            results_sheet.title = 'Test natijalari'
            
            headers = [
                "O'quvchi ismi",
                "Sinf",
                "Jami savollar",
                "To'g'ri javoblar",
                "Noto'g'ri javoblar",
                "Ball",
                "Foiz",
                "Baholash",
                "Qayta ishlangan vaqt"
            ]
            results_sheet.append(headers)
            for cell in results_sheet[1]:
                cell.font = self.header_font
            
            for result in results:
                results_sheet.append([
                    result.student_name,
                    result.student_class or '',
                    result.total_questions,
                    result.correct_answers,
                    result.wrong_answers,
                    result.score,
                    f"{result.percentage:.1f}%",
                    result.grade,
                    result.processed_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            
            self._auto_fit_columns(results_sheet)
            
            summary_sheet = workbook.create_sheet("Umumiy ma'lumot")
            summary_data = [
                ("Test nomi", test.title),
                ("Fan", test.get_subject_display()),
                ("Sinf darajasi", test.grade_level),
                ("Jami o'quvchilar", len(results)),
                ("O'rtacha foiz", f"{sum(r.percentage for r in results) / len(results):.1f}%"),
                ("Eksport vaqti", results[0].processed_at.strftime('%Y-%m-%d %H:%M:%S'))
            ]
            summary_sheet.append(["Ko'rsatkich", "Qiymat"])
            for cell in summary_sheet[1]:
                cell.font = self.header_font
            
            for label, value in summary_data:
                summary_sheet.append([label, value])
            
            self._auto_fit_columns(summary_sheet)
            
            workbook.save(excel_path)
            return excel_path
            
        except Exception as e:
            logger.error(f"Excel eksportda xatolik: {e}")
            return None
